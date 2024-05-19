import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('May_AQI.xlsx')  # 開啟 Excel 檔案
wsec1 = openpyxl.Workbook()  # 開啟 Excel 檔案

s1 = wb['Sheet']  # 取得工作表名稱為「工作表1」的內容


def get_values(sheet):
    arr = []  # 第一層串列
    for row in sheet:
        arr2 = []  # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr


data_json_list_chinese = ['測站編號', '測站名稱', '監測日期', '空氣品質指標', '二氧化硫副指標', '一氧化碳副指標',
                          '臭氧副指標', '懸浮微粒副指標',
                          '二氧化氮副指標', '臭氧8小時副指標', '細懸浮微粒副指標']
now = datetime.now().strftime("%Y-%m-%d")
print(type(now))
day = [str(i).zfill(2) for i in range(1, 32)]
for z in day:
    date = f"2024-05-{z}"
    # print(date)
    new_sheet = wsec1.create_sheet(date)  # 在 Workbook 物件上創建新的工作表
    new_sheet.append(data_json_list_chinese)  # 寫入標題
    for a in get_values(s1):
        if a == data_json_list_chinese:
            continue
        else:
            if a[2] == now:
                break
            elif a[2] == date:
                new_sheet.append(a)  # 寫入標題
wsec1.remove(wsec1['Sheet'])  # 刪除預設工作表
wsec1.save('New_May_AQI.xlsx')  # 儲存到新的檔案
