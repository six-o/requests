import requests
from bs4 import BeautifulSoup
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

api_url = 'https://data.epa.gov.tw/api/v2'
dataset = "AQX_P_434"  # 請替換欲下載資料集代碼,如何查詢資料集代碼詳簡報第8頁
Format = "json"  # 平臺提供格式包含JSON、 XML、 CSV，請依欲下載資料格式自行更換。
offset = 0  # 為遞移起始下載筆數使用，請輸入欲跳過的筆數，若欲擷取第1001筆後之資料，數字請輸入 1000
# 將項目數量放入變數 limit
limit = 858  # 請輸入欲取得資料的筆數，考量系統效能，資料擷取上限為 1000 筆,若未設定，預設值為 1000
api_key = "b03125d5-9a0d-4d4a-a88d-44085566fb4a"  # 加入會員後可取得 api_key，請自行更換。
API_URL = f'{api_url}/{dataset}?format={Format}&offset={offset}&limit={limit}&api_key={api_key}'

data = requests.get(API_URL)  # 使用 get 方法透過空氣品質指標 API 取得內容
data_json = data.json()  # 將取得的檔案轉換為 JSON 格式

data_json_list_english = ['siteid', 'sitename', 'monitordate', 'aqi', 'so2subindex', 'cosubindex', 'o3subindex',
                          'pm10subindex',
                          'no2subindex', 'o38subindex', 'pm25subindex']

data_json_list_chinese = ['測站編號', '測站名稱', '監測日期', '空氣品質指標', '二氧化硫副指標', '一氧化碳副指標',
                          '臭氧副指標', '懸浮微粒副指標',
                          '二氧化氮副指標', '臭氧8小時副指標', '細懸浮微粒副指標']

vb = openpyxl.Workbook()  # 建立一個新的工作簿
vs = vb.active  # 取得目前工作表

for i, header in enumerate(data_json_list_chinese):  # 列印標題
    vs.cell(row=1, column=i + 1, value=header)
for i, record in enumerate(data_json['records'], start=2):  # 逐筆取出資料
    for j, key in enumerate(data_json_list_english):
        if key == 'monitordate':
            vs.cell(row=i, column=j + 1, value=str(record[key]))
        else:
            vs.cell(row=i, column=j + 1, value=record[key] if key in record else '')
vb.save('May_AQI.xlsx')  # 將工作簿存檔

# 文字檔輸出
# with open('test.txt', 'w', encoding='utf-8') as f:
#     for a in data_json_list_chinese: # 列印標題
#         f.write(f'{a}\t')
#     f.write('\n')
#     for j in data_json['records']: # 逐筆取出資料
#         for k in data_json_list_english:
#             f.write(f'{j[k]}\t')
#         print(f'{j['sitename']}', end='\t')
#         f.write('\n')

# Fake

# 將獲取的數據轉換為 pandas DataFrame
# df = pd.DataFrame(data_json['records'])
#
# # 將 'monitordate' 列轉換為 datetime 對象，並設置為索引
# df['monitordate'] = pd.to_datetime(df['monitordate'])
# df = df.set_index('monitordate')
#
# # 將欄位轉換為數字
# for col in ['aqi', 'so2subindex', 'cosubindex', 'o3subindex', 'pm10subindex', 'no2subindex', 'o38subindex', 'pm25subindex']:
#     df[col] = pd.to_numeric(df[col], errors='coerce')
#
# # 現在計算平均值
# df_daily = df.resample('D').mean()
# # 繪製折線圖
# df[['aqi', 'so2subindex', 'cosubindex', 'o3subindex', 'pm10subindex', 'no2subindex', 'o38subindex', 'pm25subindex']].plot(kind='line')
# plt.title('不同指標隨時間變化的趨勢')
# plt.xlabel('日期')
# plt.ylabel('值')
# plt.show()
#
# # 繪製長條圖
# df_daily[['aqi', 'so2subindex', 'cosubindex', 'o3subindex', 'pm10subindex', 'no2subindex', 'o38subindex', 'pm25subindex']].mean().plot(kind='bar')
# plt.title('不同指標的比較')
# plt.xlabel('指標')
# plt.ylabel('平均值')
# plt.show()
#
# # 繪製 AQI 圖表
# df_daily['aqi'].plot(kind='line')
# plt.title('AQI隨時間變化')
# plt.xlabel('日期')
# plt.ylabel('AQI')
# plt.show()